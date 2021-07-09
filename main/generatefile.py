from main.models import excel_generation_request
from django.http.response import FileResponse
from openpyxl import Workbook
from django.http import FileResponse


def GenerateExcelFile(data, id):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = data.get('alpha3Code')
    country_name = data.get('name')
    ws['B1'] = (f'Monthly data for {country_name}')
    ws['B3'] = ('Beneficiaries assisted (Male):')
    ws['B4'] = ('Beneficiaries assisted (Female):')
    currency = data.get('currency'[0])
    ws['B5'] = (f'Cash value of assistance {currency}:')

    file_name = f"{country_name}-{id}.xlsx"

    wb.save(file_name)

    obj = excel_generation_request.objects.get(id = id)
    obj.generated_file = file_name
    obj.save()